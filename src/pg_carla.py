#!/usr/bin/env python

# Copyright (c) 2019: Jianyu Chen (jianyuchen@berkeley.edu).
#
# This work is licensed under the terms of the MIT license.
# For a copy, see <https://opensource.org/licenses/MIT>.

import gym
import gym_carla
import carla
import pyglet
import time
import matplotlib.pyplot as plt

from carla import ColorConverter as cc

import numpy as np
import array
import argparse
import collections
import datetime
import logging
import math
import random
import re
import weakref
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

from pg_carla_agent import *


def writexslx(v, a, p):
    # print (vehicle)
    df = pd.DataFrame({"v": v, "a": a, "p": p})
    writer = pd.ExcelWriter("data.xlsx")
    df.to_excel(writer, index=False)
    writer.save()


data_folder = Path("../data")
file_to_open = data_folder / "baseline.xls"
# file_to_open = 'baseline.xls'

sh_name = "set"
excel_data_df = pd.read_excel(file_to_open, sheet_name=sh_name)
# wb = load_workbook(file_to_open)
# print(wb.sheetnames)
# ws = wb[sh_name]
#
vs = excel_data_df["vs"].tolist()
vm = excel_data_df["vm"].tolist()
a = excel_data_df["a"].tolist()
ps = excel_data_df["p"].tolist()

zipped_tvset = [item for item in enumerate(vs)]
[tset, vset] = zip(*zipped_tvset)
zipped_tvmset = [item for item in enumerate(vm)]
[tmea, vmea] = zip(*zipped_tvmset)


def main():
    # parameters for the gym_carla environment
    params = {
        "number_of_vehicles": 0,
        "number_of_walkers": 0,
        "display_size": 256,  # screen size of bird-eye render
        "max_past_step": 1,  # the number of past steps to draw
        "dt": 0.1,  # time interval between two frames
        "discrete": False,  # whether to use discrete control space
        "discrete_acc": [-3.0, 0.0, 3.0],  # discrete value of accelerations
        "discrete_steer": [-0.2, 0.0, 0.2],  # discrete value of steering angles
        "continuous_accel_range": [-3.0, 3.0],  # continuous acceleration range
        "continuous_steer_range": [-0.3, 0.3],  # continuous steering angle range
        "ego_vehicle_filter": "vehicle.lincoln*",  # filter for defining ego vehicle
        "carlaserver": "192.168.60.80",  # carla server ip address
        "port": 2000,  # connection port
        "town": "Town03",  # which town to simulate
        "task_mode": "random",  # mode of the task, [random, roundabout (only for Town03)]
        "max_time_episode": 1000,  # maximum timesteps per episode
        "max_waypt": 12,  # maximum number of waypoints
        "obs_range": 32,  # observation range (meter)
        "lidar_bin": 0.125,  # bin size of lidar sensor (meter)
        "d_behind": 12,  # distance behind the ego vehicle (meter)
        "out_lane_thres": 2.0,  # threshold for out of lane
        "desired_speed": 40,  # desired speed (m/s)
        "max_ego_spawn_times": 200,  # maximum times to spawn ego vehicle
        "display_route": True,  # whether to render the desired route
        "pixor_size": 64,  # size of the pixor labels
        "pixor": False,  # whether to output PIXOR observation
        "file_path": "../data/straight.xodr",
    }

    # Set gym-carla environment
    env = gym.make("carla-v0", params=params)

    print("simulation starts")
    xs, hs, dlogps, drs = [], [], [], []
    acts = []
    running_reward = None
    reward_sum = 0
    episode_number = 0
    counter = 0

    velocity, acceleration, pedal, timing = [], [], [], []

    obs = env.reset()
    start = time.time()
    epi_start = start

    acc = obs["state"][4]
    vel = obs["state"][2]
    x = [vel, acc]
    # prev_x = None  #  reserved for computing the difference frame
    velocity.append(vel)
    acceleration.append(acc)
    timing.append(0)

    while True:
        while counter < 20:
            aprob, h = policy_forward(x)
            # action = [2.0, 0.0]
            action = np.random.choice(A, 1, aprob.tolist())[0]

            acts.append(action)  # logging the actions taken for visualization
            # record various intermediates (needed later for backprop)
            xs.append(x)  # observation
            hs.append(h)  # hidden state
            # y = 1 if action == 2 else 0 # a "fake label"
            y = np.zeros(A)
            y[action] = 1
            # grad that encourages the action that was taken to be taken
            dlogps.append(y - aprob)
            # (see http://cs231n.github.io/neural-networks-2/#losses if confused)

            # step the environment and get new measurements
            action = [throttle_space[action], 0]
            obs, r, done, info = env.step(action)

            # print("-----------------------------")
            time_instant = time.time()
            sec = int(time_instant - start) + counter
            vel = obs["state"][2]
            acc = obs["state"][4]
            vel_interp = np.interp(sec, tset, vset)
            r_speed = -np.sqrt(np.fabs(vel - vel_interp))
            r_egy_csm = -(acc ** 2)
            reward = 1000 * r_speed + r_egy_csm
            reward_sum += reward

            x = [vel, acc]

            velocity.append(vel)
            acceleration.append(acc)
            timing.append(time_instant - epi_start)

            # record reward (has to be done after we call step() to get reward for
            # previous action)
            drs.append(reward)

            if time_instant - start >= 1:
                start = time_instant
                counter = counter + 1

        ###### Episode done ######
        episode_number = episode_number + 1
        counter = 0

        episode_len = len(timing)
        vs_epi = np.interp(timing, tset, vset)
        vm_epi = np.interp(timing, tmea, vmea)
        plt.plot(timing, velocity, 'r-', timing, vs_epi, 'g--', timing, vm_epi, 'b-.')
        plt.show(block=False)
        # stack together all inputs, hidden states, action gradients, and
        # rewards for this episode

        epx = np.vstack(xs)
        eph = np.vstack(hs)
        epdlogp = np.vstack(dlogps)
        epr = np.vstack(drs)
        xs, hs, dlogps, drs = [], [], [], []  # reset array memory
        acts = []  # could be plotted to compare with human throttle input
        velocity, acceleration, pedal, timing = [], [], [], []

        # compute the discounted reward backwards through time
        discounted_epr = discount_rewards(epr)
        # standardize the rewards to be unit normal (helps control the gradient
        # estimator variance)
        discounted_epr -= np.mean(discounted_epr)
        discounted_epr /= np.std(discounted_epr)

        # modulate the gradient with advantage (PG magic happens right here.)
        epdlogp *= discounted_epr
        grad = policy_backward(epx, eph, epdlogp)
        for k in model:
            grad_buffer[k] += grad[k]  # accumulate grad over batch

        if episode_number % batch_size == 0:
            for k, v in list(model.items()):
                g = grad_buffer[k]  # gradient
                rmsprop_cache[k] = (
                    decay_rate * rmsprop_cache[k] + (1 - decay_rate) * g ** 2
                )
                model[k] += learning_rate * g / (np.sqrt(rmsprop_cache[k]) + 1e-5)
                # reset batch gradient buffer
                grad_buffer[k] = np.zeros_like(v)

        # boring book-keeping
        running_reward = (
            reward_sum
            if running_reward is None
            else running_reward * 0.99 + reward_sum * 0.01
        )
        print(
            "resetting env. episode %d reward total was %f. running mean: %f"
            % (episode_number, reward_sum, running_reward)
        )

        reward_sum = 0
        if episode_number % 100 == 0:
            pickle.dump(model, open("veos-save2.p", "wb"))

        obs = env.reset()  # reset env
        start = time.time()
        epi_start = start
        vel = obs["state"][2]
        acc = obs["state"][4]
        x = [vel, acc]
        velocity.append(vel)
        acceleration.append(acc)
        timing.append(0)
        # prev_x = None  #  reserved for computing the difference frame


if __name__ == "__main__":
    main()
